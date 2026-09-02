"""
================================================================================
 FUNDING RATE ARBITRAGE SCANNER  (ccxt version - for GitHub Actions / PC)
================================================================================
Fetches HISTORICAL FUNDING RATES for ALL altcoin perpetual contracts
(stablecoins and BTC excluded) across:

    Binance, Bybit, KuCoin, MEXC, Kraken, HTX, Gate.Io, Coinbase,
    Hyperliquid(DEX), Bitget

...over the last 200 days, using ccxt's unified API. All exchanges are
fetched IN PARALLEL (separate threads). If a coin fails, it is retried
(with backoff) on the SAME coin before moving on.

Produces an Excel report laid out like:

  SL No | Coin Name | Date & Time | Spread | Binance | Bybit | KuCoin | ...

...one row per coin per funding timestamp, one column per exchange, with
the highest and lowest rate in each row shown in BOLD.

--------------------------------------------------------------------------
SETUP
--------------------------------------------------------------------------
    pip install -r requirements.txt

RUN:
    python funding_arbitrage_scanner.py test      # ETH only, quick check
    python funding_arbitrage_scanner.py fetch     # full fetch, resumable
    python funding_arbitrage_scanner.py report    # build Excel from CSV so far
    python funding_arbitrage_scanner.py all       # fetch + report

Safe to stop (Ctrl+C) any time - `funding_checkpoint.json` remembers which
(exchange, symbol) pairs are already done, so re-running "fetch" resumes.

Files created next to this script:
    funding_checkpoint.json        - resume state
    funding_raw_all.csv            - all funding-rate rows collected
    funding_errors.log             - symbols that failed even after retries
    funding_arbitrage_report.xlsx  - final report
================================================================================
"""

import os
import sys
import json
import time
import csv
import threading
from datetime import datetime, timedelta, timezone
from concurrent.futures import ThreadPoolExecutor

import ccxt

# --------------------------------------------------------------------------
# CONFIG
# --------------------------------------------------------------------------
DAYS_BACK = 200
NOW_MS = int(time.time() * 1000)
START_MS = NOW_MS - DAYS_BACK * 24 * 60 * 60 * 1000
IST = timezone(timedelta(hours=5, minutes=30))

# ALTCOIN-ONLY FILTER: stablecoins and BTC excluded.
# (Remove "BTC" from this set if you also want BTC included.)
STABLECOINS = {
    "USDT", "USDC", "DAI", "BUSD", "TUSD", "USDD", "USDE", "FDUSD", "PYUSD",
    "GUSD", "USDP", "EURT", "EUR", "USTC", "UST", "FRAX", "LUSD", "SUSD",
    "USD1", "XUSD",
}
EXCLUDE_COINS = STABLECOINS | {"BTC"}

MAX_RETRIES = 5           # per-symbol retries before giving up on that symbol
RETRY_BASE_DELAY = 3      # seconds, multiplied by attempt number (backoff)
PAGE_LIMIT = 500          # rows per ccxt fetchFundingRateHistory call

HERE = os.path.dirname(os.path.abspath(__file__))
CHECKPOINT_FILE = os.path.join(HERE, "funding_checkpoint.json")
RAW_CSV_FILE = os.path.join(HERE, "funding_raw_all.csv")
ERROR_LOG_FILE = os.path.join(HERE, "funding_errors.log")
REPORT_FILE = os.path.join(HERE, "funding_arbitrage_report.xlsx")

CSV_FIELDS = ["exchange", "symbol", "coin", "funding_time_utc",
              "funding_time_ist", "funding_rate_pct"]

csv_lock = threading.Lock()
checkpoint_lock = threading.Lock()
log_lock = threading.Lock()

# Display name -> ccxt exchange id. Display name + this exact order is also
# used as the column order in the Excel report.
EXCHANGE_IDS = {
    "Binance": "binanceusdm",
    "Bybit": "bybit",
    "KuCoin": "kucoinfutures",
    "MEXC": "mexc",
    "Kraken": "krakenfutures",
    "HTX": "htx",
    "Gate.Io": "gateio",
    "Coinbase": "coinbaseinternational",
    "Hyperliquid(DEX)": "hyperliquid",
    "Bitget": "bitget",
}
EXCHANGE_ORDER = list(EXCHANGE_IDS.keys())


def log_error(msg):
    print("  [ERROR] " + msg)
    with log_lock:
        with open(ERROR_LOG_FILE, "a", encoding="utf-8") as f:
            f.write(f"{datetime.now().isoformat()}  {msg}\n")


def to_ist_str(ms):
    return datetime.fromtimestamp(ms / 1000, tz=timezone.utc).astimezone(IST).strftime("%Y-%m-%d %H:%M")


def to_utc_str(ms):
    return datetime.fromtimestamp(ms / 1000, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def load_json(path, default):
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return default
    return default


def save_json_locked(path, data, lock):
    with lock:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f)


def append_rows_csv(rows):
    if not rows:
        return
    with csv_lock:
        file_exists = os.path.exists(RAW_CSV_FILE)
        with open(RAW_CSV_FILE, "a", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=CSV_FIELDS)
            if not file_exists:
                w.writeheader()
            for r in rows:
                w.writerow(r)


def base_coin_of(market):
    return (market.get("base") or "").upper()


def get_perp_markets(exchange):
    """Return list of (unified_symbol, base_coin) for altcoin perpetuals."""
    markets = exchange.load_markets()
    out = []
    for sym, m in markets.items():
        if not m.get("swap"):
            continue
        if m.get("expiry"):
            continue
        base = base_coin_of(m)
        if base in EXCLUDE_COINS:
            continue
        quote = (m.get("quote") or "").upper()
        if quote not in ("USDT", "USDC", "USD"):
            continue
        out.append((sym, base))
    return out


def fetch_symbol_history_with_retry(exchange, ex_name, symbol, coin):
    """Fetch full funding history for one symbol, retrying the SAME symbol
    on failure (bounded) instead of skipping straight to the next coin."""
    all_rows = []
    since = START_MS
    attempt = 0
    while since < NOW_MS:
        try:
            batch = exchange.fetch_funding_rate_history(symbol, since=since, limit=PAGE_LIMIT)
            attempt = 0
            if not batch:
                break
            for entry in batch:
                t = entry.get("timestamp")
                rate = entry.get("fundingRate")
                if t is None or rate is None:
                    continue
                if t < START_MS or t > NOW_MS:
                    continue
                all_rows.append({
                    "exchange": ex_name,
                    "symbol": symbol,
                    "coin": coin,
                    "funding_time_utc": to_utc_str(t),
                    "funding_time_ist": to_ist_str(t),
                    "funding_rate_pct": round(rate * 100, 6),
                })
            newest = max(e["timestamp"] for e in batch if e.get("timestamp"))
            if newest <= since or len(batch) < PAGE_LIMIT:
                break
            since = newest + 1
            time.sleep(exchange.rateLimit / 1000)
        except Exception as e:
            attempt += 1
            if attempt > MAX_RETRIES:
                log_error(f"{ex_name} {symbol}: gave up after {MAX_RETRIES} retries ({e})")
                break
            wait = RETRY_BASE_DELAY * attempt
            print(f"  [retry {attempt}/{MAX_RETRIES}] {ex_name} {symbol} failed ({e}); retrying in {wait}s")
            time.sleep(wait)
    return all_rows


def run_exchange(ex_name, checkpoint, test_mode):
    print(f"\n=== {ex_name} starting ===")
    try:
        klass = getattr(ccxt, EXCHANGE_IDS[ex_name])
        exchange = klass({"enableRateLimit": True})
    except Exception as e:
        log_error(f"{ex_name}: could not initialize ccxt exchange ({e})")
        return

    try:
        markets = get_perp_markets(exchange)
    except Exception as e:
        log_error(f"{ex_name}: could not load markets ({e})")
        return

    if test_mode:
        markets = [(s, c) for s, c in markets if c == "ETH"][:1]

    print(f"  {ex_name}: {len(markets)} altcoin perpetual symbols")

    for i, (symbol, coin) in enumerate(markets, 1):
        key = f"{ex_name}:{symbol}"
        if not test_mode and checkpoint.get(key) == "done":
            continue
        rows = fetch_symbol_history_with_retry(exchange, ex_name, symbol, coin)
        append_rows_csv(rows)
        print(f"  [{ex_name} {i}/{len(markets)}] {symbol}: {len(rows)} records")
        if not test_mode:
            checkpoint[key] = "done"
            save_json_locked(CHECKPOINT_FILE, checkpoint, checkpoint_lock)

    print(f"=== {ex_name} finished ===")


def fetch_all(test_mode=False):
    checkpoint = load_json(CHECKPOINT_FILE, {})
    with ThreadPoolExecutor(max_workers=len(EXCHANGE_IDS)) as pool:
        futures = [pool.submit(run_exchange, name, checkpoint, test_mode)
                   for name in EXCHANGE_IDS]
        for f in futures:
            f.result()
    print("\nAll exchanges done. Raw data:", RAW_CSV_FILE)


# --------------------------------------------------------------------------
# REPORT (Excel) - pivoted layout matching:
#   SL No | Coin Name | Date & Time | Spread | Binance | Bybit | KuCoin | ...
# with the max & min rate in each row shown in BOLD.
# --------------------------------------------------------------------------
def build_report():
    import pandas as pd
    from openpyxl.styles import Font, PatternFill

    if not os.path.exists(RAW_CSV_FILE):
        print("No raw data found yet. Run 'fetch' first.")
        return

    df = pd.read_csv(RAW_CSV_FILE)
    if df.empty:
        print("Raw data file is empty.")
        return

    df["funding_time_utc_dt"] = pd.to_datetime(df["funding_time_utc"])
    df["window_utc"] = df["funding_time_utc_dt"].dt.round("h")
    df["window_ist"] = (df["window_utc"] + pd.Timedelta(hours=5, minutes=30)).dt.strftime("%d-%m-%Y %H:%M")
    df["coin_pair"] = df["coin"] + "USDT"

    # one row per (coin, timestamp), one column per exchange
    pivot = df.pivot_table(index=["coin_pair", "window_ist", "window_utc"],
                            columns="exchange", values="funding_rate_pct", aggfunc="first")

    # keep exchange columns in the fixed display order, even if some are missing
    for ex in EXCHANGE_ORDER:
        if ex not in pivot.columns:
            pivot[ex] = None
    pivot = pivot[EXCHANGE_ORDER]

    pivot["Spread"] = pivot[EXCHANGE_ORDER].max(axis=1, skipna=True) - pivot[EXCHANGE_ORDER].min(axis=1, skipna=True)
    pivot = pivot.reset_index().sort_values(["window_utc", "coin_pair"]).reset_index(drop=True)
    pivot = pivot[pivot["Spread"].notna() & (pivot["Spread"] > 0)]
    pivot.insert(0, "SL No", range(1, len(pivot) + 1))
    pivot = pivot.rename(columns={"coin_pair": "Coin Name", "window_ist": "Date & Time"})
    pivot = pivot.drop(columns=["window_utc"])

    final_cols = ["SL No", "Coin Name", "Date & Time", "Spread"] + EXCHANGE_ORDER
    pivot = pivot[final_cols]

    with pd.ExcelWriter(REPORT_FILE, engine="openpyxl") as writer:
        pivot.to_excel(writer, sheet_name="Funding_Report", index=False)
        df.drop(columns=["funding_time_utc_dt", "window_utc", "coin_pair"]).to_excel(
            writer, sheet_name="Raw_Data", index=False)

        wb = writer.book
        ws = wb["Funding_Report"]
        bold = Font(bold=True)
        header_fill = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")
        for cell in ws[1]:
            cell.font = bold
            cell.fill = header_fill
        ws.freeze_panes = "E2"

        ex_col_start = 5  # column E = first exchange column (Binance)
        ex_col_end = ex_col_start + len(EXCHANGE_ORDER) - 1
        for row in range(2, ws.max_row + 1):
            vals = []
            for col in range(ex_col_start, ex_col_end + 1):
                v = ws.cell(row=row, column=col).value
                if v is not None:
                    vals.append((col, v))
            if len(vals) < 2:
                continue
            max_col = max(vals, key=lambda x: x[1])[0]
            min_col = min(vals, key=lambda x: x[1])[0]
            ws.cell(row=row, column=max_col).font = bold
            ws.cell(row=row, column=min_col).font = bold
            ws.cell(row=row, column=4).font = bold  # Spread column always bold

    print(f"Report written: {REPORT_FILE}")
    print(f"  {len(pivot)} rows across {df['coin'].nunique()} altcoins")


# --------------------------------------------------------------------------
if __name__ == "__main__":
    mode = sys.argv[1] if len(sys.argv) > 1 else "all"
    if mode == "test":
        fetch_all(test_mode=True)
    elif mode == "fetch":
        fetch_all(test_mode=False)
    elif mode == "report":
        build_report()
    elif mode == "all":
        fetch_all(test_mode=False)
        build_report()
    else:
        print("Usage: python funding_arbitrage_scanner.py [test|fetch|report|all]")
